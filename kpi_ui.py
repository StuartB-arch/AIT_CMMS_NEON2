"""
KPI Management UI for Managers
Provides dashboard, data input, and export capabilities
"""

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from kpi_manager import KPIManager
from datetime import datetime
import traceback
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class KPIDashboard(QWidget):
    """Main KPI Dashboard for Managers"""

    def __init__(self, pool, current_user, parent=None):
        super().__init__(parent)
        self.pool = pool
        self.current_user = current_user
        self.kpi_manager = KPIManager(pool)
        self.current_period = datetime.now().strftime('%Y-%m')
        self.init_ui()

    def init_ui(self):
        """Initialize the user interface"""
        layout = QVBoxLayout()

        # Title
        title_label = QLabel("2025 KPI Dashboard - Manager View")
        title_label.setStyleSheet("font-size: 18pt; font-weight: bold; color: #2c3e50;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # Period selector
        period_layout = QHBoxLayout()
        period_layout.addWidget(QLabel("Measurement Period:"))

        self.period_combo = QComboBox()
        self.populate_periods()
        self.period_combo.currentTextChanged.connect(self.on_period_changed)
        period_layout.addWidget(self.period_combo)

        period_layout.addStretch()

        # Refresh button
        refresh_btn = QPushButton("🔄 Refresh Data")
        refresh_btn.clicked.connect(self.refresh_dashboard)
        period_layout.addWidget(refresh_btn)

        # Calculate Auto KPIs button
        calc_btn = QPushButton("📊 Calculate Auto KPIs")
        calc_btn.clicked.connect(self.calculate_auto_kpis)
        calc_btn.setStyleSheet("background-color: #3498db; color: white; font-weight: bold; padding: 8px;")
        period_layout.addWidget(calc_btn)

        layout.addLayout(period_layout)

        # Tab widget for different sections
        tab_widget = QTabWidget()

        # Tab 1: KPI Overview
        self.overview_tab = self.create_overview_tab()
        tab_widget.addTab(self.overview_tab, "📈 KPI Overview")

        # Tab 2: Manual Data Input
        self.input_tab = self.create_input_tab()
        tab_widget.addTab(self.input_tab, "📝 Manual Data Input")

        # Tab 3: Export
        self.export_tab = self.create_export_tab()
        tab_widget.addTab(self.export_tab, "📄 Export Reports")

        layout.addWidget(tab_widget)

        self.setLayout(layout)
        self.refresh_dashboard()

    def populate_periods(self):
        """Populate period dropdown with last 12 months"""
        self.period_combo.clear()
        current = datetime.now()

        for i in range(12):
            month = current.month - i
            year = current.year

            while month < 1:
                month += 12
                year -= 1

            period = f"{year}-{month:02d}"
            display = datetime(year, month, 1).strftime("%B %Y")
            self.period_combo.addItem(display, period)

    def on_period_changed(self, text):
        """Handle period selection change"""
        self.current_period = self.period_combo.currentData()
        self.refresh_dashboard()

    def create_overview_tab(self):
        """Create KPI overview tab"""
        widget = QWidget()
        layout = QVBoxLayout()

        # Summary cards
        summary_layout = QHBoxLayout()

        self.total_kpis_label = QLabel("Total KPIs: 0")
        self.total_kpis_label.setStyleSheet(self.get_card_style("#3498db"))
        summary_layout.addWidget(self.total_kpis_label)

        self.passing_kpis_label = QLabel("Passing: 0")
        self.passing_kpis_label.setStyleSheet(self.get_card_style("#27ae60"))
        summary_layout.addWidget(self.passing_kpis_label)

        self.failing_kpis_label = QLabel("Failing: 0")
        self.failing_kpis_label.setStyleSheet(self.get_card_style("#e74c3c"))
        summary_layout.addWidget(self.failing_kpis_label)

        self.pending_kpis_label = QLabel("Pending Data: 0")
        self.pending_kpis_label.setStyleSheet(self.get_card_style("#f39c12"))
        summary_layout.addWidget(self.pending_kpis_label)

        layout.addLayout(summary_layout)

        # KPI Results Table
        self.overview_table = QTableWidget()
        self.overview_table.setColumnCount(7)
        self.overview_table.setHorizontalHeaderLabels([
            "Function", "KPI Name", "Value", "Target", "Status", "Calculated Date", "Notes"
        ])
        self.overview_table.horizontalHeader().setStretchLastSection(True)
        self.overview_table.setAlternatingRowColors(True)
        self.overview_table.setSelectionBehavior(QTableWidget.SelectRows)

        layout.addWidget(self.overview_table)

        widget.setLayout(layout)
        return widget

    def create_input_tab(self):
        """Create manual data input tab"""
        widget = QWidget()
        layout = QVBoxLayout()

        # Instructions
        instructions = QLabel(
            "Select a KPI below to enter manual data. Some KPIs are calculated automatically from the database."
        )
        instructions.setWordWrap(True)
        instructions.setStyleSheet("background-color: #e8f4f8; padding: 10px; border-radius: 5px;")
        layout.addWidget(instructions)

        # KPI Selection
        selection_layout = QHBoxLayout()
        selection_layout.addWidget(QLabel("Select KPI:"))

        self.kpi_selector = QComboBox()
        self.populate_kpi_selector()
        self.kpi_selector.currentTextChanged.connect(self.on_kpi_selected)
        selection_layout.addWidget(self.kpi_selector, 1)

        layout.addLayout(selection_layout)

        # Scroll area for input fields
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMinimumHeight(400)

        self.input_form_widget = QWidget()
        self.input_form_layout = QVBoxLayout()
        self.input_form_widget.setLayout(self.input_form_layout)
        scroll.setWidget(self.input_form_widget)

        layout.addWidget(scroll)

        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        save_btn = QPushButton("💾 Save Data")
        save_btn.clicked.connect(self.save_manual_data)
        save_btn.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; padding: 10px 20px;")
        button_layout.addWidget(save_btn)

        calc_btn = QPushButton("🧮 Calculate KPI")
        calc_btn.clicked.connect(self.calculate_selected_kpi)
        calc_btn.setStyleSheet("background-color: #3498db; color: white; font-weight: bold; padding: 10px 20px;")
        button_layout.addWidget(calc_btn)

        layout.addLayout(button_layout)

        widget.setLayout(layout)
        return widget

    def create_export_tab(self):
        """Create export tab"""
        widget = QWidget()
        layout = QVBoxLayout()

        # Export options
        layout.addWidget(QLabel("<h3>Export KPI Reports</h3>"))

        info_label = QLabel(
            "Export your KPI data to professional PDF or Excel reports for presentation and archival."
        )
        info_label.setWordWrap(True)
        info_label.setStyleSheet("background-color: #e8f4f8; padding: 10px; border-radius: 5px;")
        layout.addWidget(info_label)

        layout.addSpacing(20)

        # Period selection for export
        period_layout = QHBoxLayout()
        period_layout.addWidget(QLabel("Export Period:"))
        self.export_period_combo = QComboBox()
        self.export_period_combo.addItem("Current Period", self.current_period)
        self.export_period_combo.addItem("Last 3 Months", "3months")
        self.export_period_combo.addItem("Last 6 Months", "6months")
        self.export_period_combo.addItem("Last 12 Months", "12months")
        period_layout.addWidget(self.export_period_combo)
        period_layout.addStretch()
        layout.addLayout(period_layout)

        layout.addSpacing(20)

        # Export buttons
        export_btn_layout = QHBoxLayout()

        pdf_btn = QPushButton("📄 Export to PDF")
        pdf_btn.setIcon(self.style().standardIcon(QStyle.SP_FileIcon))
        pdf_btn.clicked.connect(self.export_to_pdf)
        pdf_btn.setStyleSheet("background-color: #e74c3c; color: white; font-weight: bold; padding: 15px; font-size: 12pt;")
        pdf_btn.setMinimumHeight(60)
        export_btn_layout.addWidget(pdf_btn)

        excel_btn = QPushButton("📊 Export to Excel")
        excel_btn.setIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView))
        excel_btn.clicked.connect(self.export_to_excel)
        excel_btn.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold; padding: 15px; font-size: 12pt;")
        excel_btn.setMinimumHeight(60)
        export_btn_layout.addWidget(excel_btn)

        layout.addLayout(export_btn_layout)

        # Export history
        layout.addSpacing(30)
        layout.addWidget(QLabel("<h3>Export History</h3>"))

        self.export_history_table = QTableWidget()
        self.export_history_table.setColumnCount(5)
        self.export_history_table.setHorizontalHeaderLabels([
            "Export Date", "Period", "Type", "Exported By", "File Name"
        ])
        self.export_history_table.horizontalHeader().setStretchLastSection(True)
        self.export_history_table.setAlternatingRowColors(True)
        layout.addWidget(self.export_history_table)

        layout.addStretch()

        widget.setLayout(layout)
        return widget

    def get_card_style(self, color):
        """Get style for summary cards"""
        return f"""
            QLabel {{
                background-color: {color};
                color: white;
                padding: 20px;
                border-radius: 10px;
                font-size: 14pt;
                font-weight: bold;
            }}
        """

    def populate_kpi_selector(self):
        """Populate KPI selector with manual KPIs"""
        self.kpi_selector.clear()
        self.kpi_selector.addItem("-- Select KPI --", None)

        manual_kpis = self.kpi_manager.get_kpis_needing_manual_data()
        for kpi_name in manual_kpis:
            self.kpi_selector.addItem(kpi_name, kpi_name)

    def on_kpi_selected(self, text):
        """Handle KPI selection for data input"""
        # Clear previous form
        while self.input_form_layout.count():
            child = self.input_form_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        kpi_name = self.kpi_selector.currentData()
        if not kpi_name:
            return

        # Get required fields
        fields = self.kpi_manager.get_required_fields_for_kpi(kpi_name)

        # Get KPI definition
        kpi_def = self.kpi_manager.get_kpi_by_name(kpi_name)

        if kpi_def:
            # Show KPI info
            info_group = QGroupBox(f"KPI: {kpi_name}")
            info_layout = QVBoxLayout()
            info_layout.addWidget(QLabel(f"<b>Description:</b> {kpi_def['description']}"))
            info_layout.addWidget(QLabel(f"<b>Formula:</b> {kpi_def['formula']}"))
            info_layout.addWidget(QLabel(f"<b>Target:</b> {kpi_def['acceptance_criteria']}"))
            info_layout.addWidget(QLabel(f"<b>Frequency:</b> {kpi_def['frequency']}"))
            info_group.setLayout(info_layout)
            self.input_form_layout.addWidget(info_group)

        # Load existing data if any
        existing_data = self.kpi_manager.get_manual_data(kpi_name, self.current_period)
        existing_dict = {row['data_field']: row['data_value'] or row['data_text'] for row in existing_data}

        # Create input fields
        inputs_group = QGroupBox("Data Input")
        inputs_layout = QFormLayout()

        self.input_fields = {}

        for field_info in fields:
            field_name = field_info['field']
            label = field_info['label']
            field_type = field_info['type']

            if field_type == 'number':
                input_widget = QDoubleSpinBox()
                input_widget.setRange(0, 999999)
                input_widget.setDecimals(2)
                input_widget.setMinimumWidth(150)
                # Load existing value
                if field_name in existing_dict and existing_dict[field_name] is not None:
                    input_widget.setValue(float(existing_dict[field_name]))
            else:  # text
                input_widget = QTextEdit()
                input_widget.setMaximumHeight(100)
                # Load existing value
                if field_name in existing_dict:
                    input_widget.setText(str(existing_dict[field_name]))

            self.input_fields[field_name] = input_widget
            inputs_layout.addRow(label + ":", input_widget)

        inputs_group.setLayout(inputs_layout)
        self.input_form_layout.addWidget(inputs_group)

        self.input_form_layout.addStretch()

    def save_manual_data(self):
        """Save manual data inputs"""
        kpi_name = self.kpi_selector.currentData()
        if not kpi_name:
            QMessageBox.warning(self, "No KPI Selected", "Please select a KPI first.")
            return

        try:
            # Save each field
            for field_name, widget in self.input_fields.items():
                if isinstance(widget, QDoubleSpinBox):
                    value = widget.value()
                    self.kpi_manager.save_manual_data(
                        kpi_name=kpi_name,
                        measurement_period=self.current_period,
                        data_field=field_name,
                        data_value=value,
                        entered_by=self.current_user
                    )
                else:  # QTextEdit
                    text = widget.toPlainText()
                    self.kpi_manager.save_manual_data(
                        kpi_name=kpi_name,
                        measurement_period=self.current_period,
                        data_field=field_name,
                        data_value=None,
                        data_text=text,
                        entered_by=self.current_user
                    )

            QMessageBox.information(self, "Success", f"Manual data saved for {kpi_name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save data: {str(e)}\n\n{traceback.format_exc()}")

    def calculate_selected_kpi(self):
        """Calculate the selected KPI from manual data"""
        kpi_name = self.kpi_selector.currentData()
        if not kpi_name:
            QMessageBox.warning(self, "No KPI Selected", "Please select a KPI first.")
            return

        try:
            result = self.kpi_manager.calculate_manual_kpi(kpi_name, self.current_period, self.current_user)

            if 'error' in result:
                QMessageBox.warning(self, "Cannot Calculate", result['error'])
                return

            msg = f"KPI Calculated: {kpi_name}\n\n"
            if result.get('value') is not None:
                msg += f"Value: {result['value']:.2f}\n"
            if result.get('text'):
                msg += f"Result: {result['text']}\n"
            if result.get('meets_criteria') is not None:
                status = "✓ PASS" if result['meets_criteria'] else "✗ FAIL"
                msg += f"Status: {status}"

            QMessageBox.information(self, "Calculation Complete", msg)
            self.refresh_dashboard()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to calculate KPI: {str(e)}\n\n{traceback.format_exc()}")

    def calculate_auto_kpis(self):
        """Calculate all automatic KPIs"""
        try:
            results = self.kpi_manager.calculate_all_auto_kpis(self.current_period, self.current_user)

            msg = "Auto KPI Calculation Complete:\n\n"
            for kpi_key, result in results.items():
                if 'error' in result:
                    msg += f"✗ {kpi_key}: {result['error']}\n"
                else:
                    msg += f"✓ {kpi_key}: Success\n"

            QMessageBox.information(self, "Calculation Complete", msg)
            self.refresh_dashboard()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to calculate auto KPIs: {str(e)}\n\n{traceback.format_exc()}")

    def refresh_dashboard(self):
        """Refresh the overview dashboard"""
        try:
            results = self.kpi_manager.get_kpi_results(self.current_period)

            # Update summary cards
            total = len(results)
            passing = sum(1 for r in results if r.get('meets_criteria') is True)
            failing = sum(1 for r in results if r.get('meets_criteria') is False)
            pending = 17 - total  # 17 total KPIs

            self.total_kpis_label.setText(f"Total KPIs: {total}")
            self.passing_kpis_label.setText(f"✓ Passing: {passing}")
            self.failing_kpis_label.setText(f"✗ Failing: {failing}")
            self.pending_kpis_label.setText(f"⏳ Pending Data: {pending}")

            # Update table
            self.overview_table.setRowCount(len(results))

            for row, result in enumerate(results):
                # Function
                self.overview_table.setItem(row, 0, QTableWidgetItem(result.get('function_code', '')))

                # KPI Name
                self.overview_table.setItem(row, 1, QTableWidgetItem(result.get('kpi_name', '')))

                # Value
                value_text = result.get('calculated_text') or (
                    f"{result['calculated_value']:.2f}" if result.get('calculated_value') is not None else 'N/A'
                )
                self.overview_table.setItem(row, 2, QTableWidgetItem(value_text))

                # Target
                self.overview_table.setItem(row, 3, QTableWidgetItem(result.get('acceptance_criteria', '')))

                # Status
                if result.get('meets_criteria') is True:
                    status_item = QTableWidgetItem("✓ PASS")
                    status_item.setBackground(QColor("#d5f4e6"))
                    status_item.setForeground(QColor("#27ae60"))
                elif result.get('meets_criteria') is False:
                    status_item = QTableWidgetItem("✗ FAIL")
                    status_item.setBackground(QColor("#fadbd8"))
                    status_item.setForeground(QColor("#e74c3c"))
                else:
                    status_item = QTableWidgetItem("N/A")
                    status_item.setBackground(QColor("#f9e79f"))

                status_item.setFont(QFont("Arial", 10, QFont.Bold))
                self.overview_table.setItem(row, 4, status_item)

                # Date
                calc_date = result.get('calculation_date')
                date_str = calc_date.strftime('%Y-%m-%d %H:%M') if calc_date else ''
                self.overview_table.setItem(row, 5, QTableWidgetItem(date_str))

                # Notes
                self.overview_table.setItem(row, 6, QTableWidgetItem(result.get('notes', '')))

            self.overview_table.resizeColumnsToContents()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to refresh dashboard: {str(e)}\n\n{traceback.format_exc()}")

    def export_to_pdf(self):
        """Export KPI data to PDF"""
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Export to PDF", f"KPI_Report_{self.current_period}.pdf", "PDF Files (*.pdf)"
        )

        if not file_name:
            return

        try:
            results = self.kpi_manager.get_kpi_results(self.current_period)

            # Create PDF
            doc = SimpleDocTemplate(file_name, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            elements.append(Paragraph("KPI Performance Report", title_style))

            # Period info
            period_text = f"Measurement Period: {self.current_period}"
            elements.append(Paragraph(period_text, styles['Normal']))
            elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
            elements.append(Paragraph(f"Generated by: {self.current_user}", styles['Normal']))
            elements.append(Spacer(1, 20))

            # Summary
            total = len(results)
            passing = sum(1 for r in results if r.get('meets_criteria') is True)
            failing = sum(1 for r in results if r.get('meets_criteria') is False)

            summary_data = [
                ['Total KPIs', 'Passing', 'Failing', 'Pending'],
                [str(total), str(passing), str(failing), str(17 - total)]
            ]

            summary_table = Table(summary_data, colWidths=[2*inch]*4)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 30))

            # KPI Details
            if results:
                elements.append(Paragraph("KPI Details", styles['Heading2']))
                elements.append(Spacer(1, 10))

                data = [['Function', 'KPI', 'Value', 'Target', 'Status']]

                for result in results:
                    value_text = result.get('calculated_text') or (
                        f"{result['calculated_value']:.2f}" if result.get('calculated_value') is not None else 'N/A'
                    )

                    if result.get('meets_criteria') is True:
                        status = "PASS"
                    elif result.get('meets_criteria') is False:
                        status = "FAIL"
                    else:
                        status = "N/A"

                    data.append([
                        result.get('function_code', ''),
                        Paragraph(result.get('kpi_name', ''), styles['Normal']),
                        Paragraph(str(value_text), styles['Normal']),
                        Paragraph(result.get('acceptance_criteria', ''), styles['Normal']),
                        status
                    ])

                kpi_table = Table(data, colWidths=[0.8*inch, 2.2*inch, 1.8*inch, 1.8*inch, 0.8*inch])
                kpi_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                elements.append(kpi_table)

            # Build PDF
            doc.build(elements)

            QMessageBox.information(self, "Success", f"PDF exported successfully to:\n{file_name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export PDF: {str(e)}\n\n{traceback.format_exc()}")

    def export_to_excel(self):
        """Export KPI data to Excel"""
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Export to Excel", f"KPI_Report_{self.current_period}.xlsx", "Excel Files (*.xlsx)"
        )

        if not file_name:
            return

        try:
            results = self.kpi_manager.get_kpi_results(self.current_period)

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "KPI Report"

            # Styles
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            ws.merge_cells('A1:G1')
            ws['A1'] = 'KPI Performance Report'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Period info
            ws['A2'] = f'Period: {self.current_period}'
            ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
            ws['A4'] = f'Generated by: {self.current_user}'

            # Headers
            row = 6
            headers = ['Function', 'KPI Name', 'Value', 'Target', 'Status', 'Calculated Date', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Data
            for result in results:
                row += 1

                value_text = result.get('calculated_text') or (
                    f"{result['calculated_value']:.2f}" if result.get('calculated_value') is not None else 'N/A'
                )

                if result.get('meets_criteria') is True:
                    status = "PASS"
                    status_fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
                elif result.get('meets_criteria') is False:
                    status = "FAIL"
                    status_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                else:
                    status = "N/A"
                    status_fill = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")

                calc_date = result.get('calculation_date')
                date_str = calc_date.strftime('%Y-%m-%d %H:%M') if calc_date else ''

                data = [
                    result.get('function_code', ''),
                    result.get('kpi_name', ''),
                    value_text,
                    result.get('acceptance_criteria', ''),
                    status,
                    date_str,
                    result.get('notes', '')
                ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col == 5:  # Status column
                        cell.fill = status_fill
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')

            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 40

            # Save
            wb.save(file_name)

            QMessageBox.information(self, "Success", f"Excel file exported successfully to:\n{file_name}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export Excel: {str(e)}\n\n{traceback.format_exc()}")
