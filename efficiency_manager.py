"""
Efficiency Tracking Manager for AIT CMMS
Author: CMMS Development Team
Date: 2025-11-13

This module provides comprehensive efficiency tracking and reporting for maintenance technicians.
Tracks labor hours, calculates efficiency rates, and generates detailed reports with visualizations.

Key Metrics:
- Annual hours per technician: 1,980 hrs
- Total technicians: 9
- Weekly availability: 342.69 hrs (all technicians)
- Target efficiency: 80%
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from database_utils import db_pool
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
from typing import Dict, List, Tuple, Optional
import calendar


class EfficiencyManager:
    """Manages efficiency tracking and reporting for maintenance technicians"""

    # Constants based on requirements
    ANNUAL_HOURS_PER_TECH = 1980  # Annual working hours per technician
    TOTAL_TECHNICIANS = 9
    WEEKLY_AVAILABILITY_HOURS = 342.69  # Total weekly hours for all technicians
    TARGET_EFFICIENCY = 0.80  # 80% target efficiency

    def __init__(self, notebook: ttk.Notebook, user_name: str):
        """
        Initialize the Efficiency Manager

        Args:
            notebook: The main application notebook to add the tab to
            user_name: Current logged-in user name
        """
        self.notebook = notebook
        self.user_name = user_name
        self.efficiency_frame = None
        self.canvas_widgets = []  # Store canvas widgets for cleanup

    def create_efficiency_tab(self):
        """Create the main Efficiency Tracking tab"""
        self.efficiency_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.efficiency_frame, text="⚡ Efficiency")

        # Main container with scrollbar
        main_container = ttk.Frame(self.efficiency_frame)
        main_container.pack(fill='both', expand=True)

        # Create canvas for scrolling
        canvas = tk.Canvas(main_container)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Pack scrollbar and canvas
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Bind mouse wheel scrolling
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)

        # === HEADER SECTION ===
        header_frame = ttk.Frame(scrollable_frame)
        header_frame.pack(fill='x', padx=10, pady=5)

        title_label = ttk.Label(
            header_frame,
            text="Technician Efficiency Tracking & Reporting",
            font=('Arial', 16, 'bold')
        )
        title_label.pack(pady=10)

        # === CONTROLS SECTION ===
        controls_frame = ttk.LabelFrame(scrollable_frame, text="Report Parameters", padding=10)
        controls_frame.pack(fill='x', padx=10, pady=5)

        # Date range selection
        date_row1 = ttk.Frame(controls_frame)
        date_row1.pack(fill='x', pady=5)

        ttk.Label(date_row1, text="Start Date:", font=('Arial', 10, 'bold')).pack(side='left', padx=5)
        self.start_date_var = tk.StringVar(value=(datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        start_date_entry = ttk.Entry(date_row1, textvariable=self.start_date_var, width=15)
        start_date_entry.pack(side='left', padx=5)

        ttk.Label(date_row1, text="End Date:", font=('Arial', 10, 'bold')).pack(side='left', padx=15)
        self.end_date_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        end_date_entry = ttk.Entry(date_row1, textvariable=self.end_date_var, width=15)
        end_date_entry.pack(side='left', padx=5)

        # Quick date range buttons
        date_row2 = ttk.Frame(controls_frame)
        date_row2.pack(fill='x', pady=5)

        ttk.Label(date_row2, text="Quick Select:", font=('Arial', 10)).pack(side='left', padx=5)
        ttk.Button(date_row2, text="Last 7 Days", command=lambda: self.set_date_range(7)).pack(side='left', padx=2)
        ttk.Button(date_row2, text="Last 30 Days", command=lambda: self.set_date_range(30)).pack(side='left', padx=2)
        ttk.Button(date_row2, text="Last 90 Days", command=lambda: self.set_date_range(90)).pack(side='left', padx=2)
        ttk.Button(date_row2, text="This Month", command=self.set_current_month).pack(side='left', padx=2)
        ttk.Button(date_row2, text="Last Month", command=self.set_last_month).pack(side='left', padx=2)
        ttk.Button(date_row2, text="Year to Date", command=self.set_year_to_date).pack(side='left', padx=2)

        # Action buttons
        action_row = ttk.Frame(controls_frame)
        action_row.pack(fill='x', pady=10)

        ttk.Button(
            action_row,
            text="🔄 Generate Report",
            command=self.generate_report,
            style='Accent.TButton'
        ).pack(side='left', padx=5)

        ttk.Button(
            action_row,
            text="🖨️ Print Report",
            command=self.print_report
        ).pack(side='left', padx=5)

        ttk.Button(
            action_row,
            text="💾 Export to Excel",
            command=self.export_to_excel
        ).pack(side='left', padx=5)

        ttk.Button(
            action_row,
            text="📄 Export to PDF",
            command=self.export_to_pdf
        ).pack(side='left', padx=5)

        # === SUMMARY METRICS SECTION ===
        self.summary_frame = ttk.LabelFrame(scrollable_frame, text="Overall Efficiency Summary", padding=10)
        self.summary_frame.pack(fill='x', padx=10, pady=5)

        # Create summary labels (will be populated when report is generated)
        self.summary_labels = {}
        summary_grid = ttk.Frame(self.summary_frame)
        summary_grid.pack(fill='x')

        # Row 1
        row1 = ttk.Frame(summary_grid)
        row1.pack(fill='x', pady=2)
        self._create_summary_metric(row1, "Total Available Hours:", "total_available", 0)
        self._create_summary_metric(row1, "Total Worked Hours:", "total_worked", 1)
        self._create_summary_metric(row1, "Overall Efficiency:", "overall_efficiency", 2)

        # Row 2
        row2 = ttk.Frame(summary_grid)
        row2.pack(fill='x', pady=2)
        self._create_summary_metric(row2, "PM Hours:", "pm_hours", 0)
        self._create_summary_metric(row2, "CM Hours:", "cm_hours", 1)
        self._create_summary_metric(row2, "Target Efficiency:", "target", 2)

        # === TECHNICIAN DETAIL TABLE ===
        self.detail_frame = ttk.LabelFrame(scrollable_frame, text="Technician Performance Details", padding=10)
        self.detail_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Create treeview for technician details
        tree_container = ttk.Frame(self.detail_frame)
        tree_container.pack(fill='both', expand=True)

        # Scrollbars for treeview
        tree_scroll_y = ttk.Scrollbar(tree_container, orient='vertical')
        tree_scroll_y.pack(side='right', fill='y')

        tree_scroll_x = ttk.Scrollbar(tree_container, orient='horizontal')
        tree_scroll_x.pack(side='bottom', fill='x')

        # Treeview columns
        columns = (
            'technician', 'available_hrs', 'pm_hrs', 'cm_hrs', 'total_hrs',
            'pm_count', 'cm_count', 'efficiency', 'vs_target', 'status'
        )

        self.tech_tree = ttk.Treeview(
            tree_container,
            columns=columns,
            show='headings',
            yscrollcommand=tree_scroll_y.set,
            xscrollcommand=tree_scroll_x.set,
            height=10
        )

        tree_scroll_y.config(command=self.tech_tree.yview)
        tree_scroll_x.config(command=self.tech_tree.xview)

        # Configure columns
        self.tech_tree.heading('technician', text='Technician')
        self.tech_tree.heading('available_hrs', text='Available Hrs')
        self.tech_tree.heading('pm_hrs', text='PM Hours')
        self.tech_tree.heading('cm_hrs', text='CM Hours')
        self.tech_tree.heading('total_hrs', text='Total Hours')
        self.tech_tree.heading('pm_count', text='PM Count')
        self.tech_tree.heading('cm_count', text='CM Count')
        self.tech_tree.heading('efficiency', text='Efficiency %')
        self.tech_tree.heading('vs_target', text='vs Target')
        self.tech_tree.heading('status', text='Status')

        self.tech_tree.column('technician', width=150)
        self.tech_tree.column('available_hrs', width=100, anchor='center')
        self.tech_tree.column('pm_hrs', width=90, anchor='center')
        self.tech_tree.column('cm_hrs', width=90, anchor='center')
        self.tech_tree.column('total_hrs', width=90, anchor='center')
        self.tech_tree.column('pm_count', width=80, anchor='center')
        self.tech_tree.column('cm_count', width=80, anchor='center')
        self.tech_tree.column('efficiency', width=100, anchor='center')
        self.tech_tree.column('vs_target', width=100, anchor='center')
        self.tech_tree.column('status', width=120, anchor='center')

        self.tech_tree.pack(fill='both', expand=True)

        # Add row coloring for status
        self.tech_tree.tag_configure('above_target', background='#d4edda')
        self.tech_tree.tag_configure('at_target', background='#fff3cd')
        self.tech_tree.tag_configure('below_target', background='#f8d7da')

        # === VISUALIZATION SECTION ===
        self.viz_frame = ttk.LabelFrame(scrollable_frame, text="Performance Visualizations", padding=10)
        self.viz_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Create notebook for different chart views
        self.chart_notebook = ttk.Notebook(self.viz_frame)
        self.chart_notebook.pack(fill='both', expand=True)

        # Initialize chart frames (charts will be populated when report is generated)
        self.chart_frames = {}
        self._initialize_chart_frames()

        # === FOOTER NOTES ===
        notes_frame = ttk.LabelFrame(scrollable_frame, text="Report Notes & Methodology", padding=10)
        notes_frame.pack(fill='x', padx=10, pady=5)

        notes_text = f"""
Efficiency Calculation Methodology:
• Annual Working Hours per Technician: {self.ANNUAL_HOURS_PER_TECH} hours
• Total Technicians: {self.TOTAL_TECHNICIANS}
• Weekly Team Availability: {self.WEEKLY_AVAILABILITY_HOURS} hours
• Target Efficiency: {self.TARGET_EFFICIENCY * 100}%

Efficiency Formula: (Total Worked Hours / Total Available Hours) × 100%
• Worked Hours = PM Hours + CM Hours (from completed work orders)
• Available Hours = (Number of working days in period) × (Weekly Availability / 5 days)

Status Indicators:
• Above Target: Efficiency ≥ 85% (Green)
• At Target: Efficiency 75-84% (Yellow)
• Below Target: Efficiency < 75% (Red)

Data Sources:
• PM Hours: pm_completions table (labor_hours + labor_minutes/60)
• CM Hours: corrective_maintenance table (labor_hours)
• Technician List: users table (active technicians only)
        """

        notes_label = ttk.Label(notes_frame, text=notes_text.strip(), justify='left', font=('Arial', 9))
        notes_label.pack(anchor='w')

        # Auto-generate report on tab creation (silently)
        self.efficiency_frame.after(500, lambda: self.generate_report(silent=True))

    def _create_summary_metric(self, parent, label_text: str, key: str, column: int):
        """Create a summary metric display"""
        metric_frame = ttk.Frame(parent)
        metric_frame.pack(side='left', expand=True, fill='x', padx=10)

        ttk.Label(metric_frame, text=label_text, font=('Arial', 9)).pack(anchor='w')
        value_label = ttk.Label(metric_frame, text="--", font=('Arial', 12, 'bold'))
        value_label.pack(anchor='w')

        self.summary_labels[key] = value_label

    def _initialize_chart_frames(self):
        """Initialize frames for different chart types"""
        chart_types = [
            ('efficiency_comparison', 'Efficiency Comparison by Technician'),
            ('hours_breakdown', 'Hours Breakdown (PM vs CM)'),
            ('workload_distribution', 'Workload Distribution'),
            ('trend_analysis', 'Efficiency Trend'),
        ]

        for key, title in chart_types:
            frame = ttk.Frame(self.chart_notebook)
            self.chart_notebook.add(frame, text=title)
            self.chart_frames[key] = frame

    def set_date_range(self, days: int):
        """Set date range to last N days"""
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)
        self.start_date_var.set(start_date.strftime('%Y-%m-%d'))
        self.end_date_var.set(end_date.strftime('%Y-%m-%d'))

    def set_current_month(self):
        """Set date range to current month"""
        today = datetime.now()
        start_date = today.replace(day=1)
        self.start_date_var.set(start_date.strftime('%Y-%m-%d'))
        self.end_date_var.set(today.strftime('%Y-%m-%d'))

    def set_last_month(self):
        """Set date range to last month"""
        today = datetime.now()
        first_of_month = today.replace(day=1)
        last_month_end = first_of_month - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        self.start_date_var.set(last_month_start.strftime('%Y-%m-%d'))
        self.end_date_var.set(last_month_end.strftime('%Y-%m-%d'))

    def set_year_to_date(self):
        """Set date range to year to date"""
        today = datetime.now()
        year_start = today.replace(month=1, day=1)
        self.start_date_var.set(year_start.strftime('%Y-%m-%d'))
        self.end_date_var.set(today.strftime('%Y-%m-%d'))

    def generate_report(self, silent=False):
        """Generate the efficiency report with all calculations and visualizations"""
        try:
            # Validate dates
            start_date = datetime.strptime(self.start_date_var.get(), '%Y-%m-%d')
            end_date = datetime.strptime(self.end_date_var.get(), '%Y-%m-%d')

            if start_date > end_date:
                if not silent:
                    messagebox.showerror("Error", "Start date must be before end date")
                return

            # Fetch data from database
            tech_data = self._fetch_technician_data(start_date, end_date)

            if not tech_data:
                if not silent:
                    messagebox.showwarning("No Data", "No efficiency data found for the selected date range")
                return

            # Calculate metrics
            self._calculate_and_display_metrics(tech_data, start_date, end_date)

            # Update technician detail table
            self._update_technician_table(tech_data)

            # Generate visualizations
            self._generate_visualizations(tech_data)

            if not silent:
                messagebox.showinfo("Success", "Efficiency report generated successfully")

        except ValueError as e:
            if not silent:
                messagebox.showerror("Error", f"Invalid date format. Use YYYY-MM-DD: {str(e)}")
        except Exception as e:
            if not silent:
                messagebox.showerror("Error", f"Failed to generate report:\n{str(e)}")
            import traceback
            print(traceback.format_exc())

    def _fetch_technician_data(self, start_date: datetime, end_date: datetime) -> List[Dict]:
        """
        Fetch technician performance data from database

        Returns list of dicts with technician data
        """
        try:
            with db_pool.get_cursor(commit=False) as cursor:
                # Get all active technicians
                cursor.execute("""
                    SELECT full_name
                    FROM users
                    WHERE is_active = TRUE AND role = 'Technician'
                    ORDER BY full_name
                """)
                technicians = [row['full_name'] for row in cursor.fetchall()]

                tech_data = []

                for tech_name in technicians:
                    # Get PM hours and count
                    # Note: dates are stored as TEXT in format YYYY-MM-DD
                    cursor.execute("""
                        SELECT
                            COUNT(*) as pm_count,
                            COALESCE(SUM(COALESCE(labor_hours, 0) + COALESCE(labor_minutes, 0)/60.0), 0) as pm_hours
                        FROM pm_completions
                        WHERE technician_name = %s
                        AND completion_date >= %s
                        AND completion_date <= %s
                    """, (tech_name, start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))

                    pm_result = cursor.fetchone()
                    pm_hours = float(pm_result['pm_hours']) if pm_result else 0.0
                    pm_count = int(pm_result['pm_count']) if pm_result else 0

                    # Debug output
                    if pm_count > 0:
                        print(f"DEBUG: {tech_name} - PM Count: {pm_count}, PM Hours: {pm_hours}")

                    # Get CM hours and count
                    # Note: Some CMs may not have closed_date even if status is 'Closed'
                    # So we check if closed_date is NULL or within range
                    cursor.execute("""
                        SELECT
                            COUNT(*) as cm_count,
                            COALESCE(SUM(COALESCE(labor_hours, 0)), 0) as cm_hours
                        FROM corrective_maintenance
                        WHERE assigned_technician = %s
                        AND status = 'Closed'
                        AND (
                            closed_date IS NULL
                            OR (closed_date >= %s AND closed_date <= %s)
                        )
                    """, (tech_name, start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))

                    cm_result = cursor.fetchone()
                    cm_hours = float(cm_result['cm_hours']) if cm_result else 0.0
                    cm_count = int(cm_result['cm_count']) if cm_result else 0

                    # Debug output
                    print(f"DEBUG: {tech_name} - CM Count: {cm_count}, CM Hours: {cm_hours}")

                    total_hours = pm_hours + cm_hours

                    tech_data.append({
                        'technician': tech_name,
                        'pm_hours': pm_hours,
                        'pm_count': pm_count,
                        'cm_hours': cm_hours,
                        'cm_count': cm_count,
                        'total_hours': total_hours
                    })

                print(f"DEBUG: Date range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
                print(f"DEBUG: Found {len(tech_data)} technicians")
                return tech_data

        except Exception as e:
            print(f"Error fetching technician data: {e}")
            import traceback
            print(traceback.format_exc())
            return []

    def _calculate_and_display_metrics(self, tech_data: List[Dict], start_date: datetime, end_date: datetime):
        """Calculate overall metrics and update summary display"""
        # Calculate number of working days in period
        days_in_period = (end_date - start_date).days + 1
        working_days = days_in_period  # Simplified - could exclude weekends if needed

        # Calculate available hours per technician for the period
        # Weekly availability per tech = 342.69 / 9 = 38.08 hrs/week
        # Daily availability per tech = 38.08 / 5 = 7.616 hrs/day
        hours_per_tech_per_day = self.WEEKLY_AVAILABILITY_HOURS / self.TOTAL_TECHNICIANS / 5
        available_hours_per_tech = hours_per_tech_per_day * working_days

        # Calculate totals
        total_available = available_hours_per_tech * len(tech_data)
        total_pm_hours = sum(t['pm_hours'] for t in tech_data)
        total_cm_hours = sum(t['cm_hours'] for t in tech_data)
        total_worked = total_pm_hours + total_cm_hours

        overall_efficiency = (total_worked / total_available * 100) if total_available > 0 else 0

        # Store available hours per tech for use in table
        for tech in tech_data:
            tech['available_hours'] = available_hours_per_tech
            tech['efficiency'] = (tech['total_hours'] / available_hours_per_tech * 100) if available_hours_per_tech > 0 else 0

        # Update summary labels
        self.summary_labels['total_available'].config(text=f"{total_available:.1f} hrs")
        self.summary_labels['total_worked'].config(text=f"{total_worked:.1f} hrs")
        self.summary_labels['pm_hours'].config(text=f"{total_pm_hours:.1f} hrs")
        self.summary_labels['cm_hours'].config(text=f"{total_cm_hours:.1f} hrs")
        self.summary_labels['target'].config(text=f"{self.TARGET_EFFICIENCY * 100:.0f}%")

        # Color code overall efficiency
        efficiency_label = self.summary_labels['overall_efficiency']
        efficiency_label.config(text=f"{overall_efficiency:.1f}%")

        if overall_efficiency >= 85:
            efficiency_label.config(foreground='green')
        elif overall_efficiency >= 75:
            efficiency_label.config(foreground='orange')
        else:
            efficiency_label.config(foreground='red')

    def _update_technician_table(self, tech_data: List[Dict]):
        """Update the technician detail treeview"""
        # Clear existing data
        for item in self.tech_tree.get_children():
            self.tech_tree.delete(item)

        # Sort by efficiency descending
        tech_data_sorted = sorted(tech_data, key=lambda x: x['efficiency'], reverse=True)

        # Insert data
        for tech in tech_data_sorted:
            efficiency = tech['efficiency']
            vs_target = efficiency - (self.TARGET_EFFICIENCY * 100)

            # Determine status and tag
            if efficiency >= 85:
                status = "Above Target"
                tag = 'above_target'
            elif efficiency >= 75:
                status = "At Target"
                tag = 'at_target'
            else:
                status = "Below Target"
                tag = 'below_target'

            self.tech_tree.insert('', 'end', values=(
                tech['technician'],
                f"{tech['available_hours']:.1f}",
                f"{tech['pm_hours']:.1f}",
                f"{tech['cm_hours']:.1f}",
                f"{tech['total_hours']:.1f}",
                tech['pm_count'],
                tech['cm_count'],
                f"{efficiency:.1f}%",
                f"{vs_target:+.1f}%",
                status
            ), tags=(tag,))

    def _generate_visualizations(self, tech_data: List[Dict]):
        """Generate all chart visualizations"""
        # Clear existing charts
        for widget in self.canvas_widgets:
            try:
                widget.destroy()
            except:
                pass
        self.canvas_widgets.clear()

        # Also clear all children in chart frames to be safe
        for frame_key, frame in self.chart_frames.items():
            for child in frame.winfo_children():
                try:
                    child.destroy()
                except:
                    pass

        # Generate each chart
        self._create_efficiency_comparison_chart(tech_data)
        self._create_hours_breakdown_chart(tech_data)
        self._create_workload_distribution_chart(tech_data)
        self._create_trend_chart(tech_data)

    def _create_efficiency_comparison_chart(self, tech_data: List[Dict]):
        """Create bar chart comparing technician efficiency"""
        frame = self.chart_frames['efficiency_comparison']

        # Sort by efficiency
        sorted_data = sorted(tech_data, key=lambda x: x['efficiency'], reverse=True)

        fig = Figure(figsize=(6, 3.5), dpi=80)
        ax = fig.add_subplot(111)

        technicians = [t['technician'] for t in sorted_data]
        efficiencies = [t['efficiency'] for t in sorted_data]

        # Color bars based on performance
        colors = []
        for eff in efficiencies:
            if eff >= 85:
                colors.append('#28a745')  # Green
            elif eff >= 75:
                colors.append('#ffc107')  # Yellow
            else:
                colors.append('#dc3545')  # Red

        bars = ax.bar(technicians, efficiencies, color=colors, alpha=0.7, edgecolor='black')

        # Add target line
        ax.axhline(y=self.TARGET_EFFICIENCY * 100, color='red', linestyle='--', linewidth=2, label=f'Target ({self.TARGET_EFFICIENCY * 100}%)')

        # Add value labels on bars
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{height:.1f}%',
                   ha='center', va='bottom', fontsize=7, fontweight='bold')

        ax.set_xlabel('Technician', fontsize=9, fontweight='bold')
        ax.set_ylabel('Efficiency (%)', fontsize=9, fontweight='bold')
        ax.set_title('Technician Efficiency Comparison', fontsize=10, fontweight='bold')
        ax.legend(fontsize=8)
        ax.grid(axis='y', alpha=0.3)

        # Rotate x labels if needed
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right', fontsize=8)
        fig.tight_layout()

        # Embed in tkinter
        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)
        self.canvas_widgets.append(canvas.get_tk_widget())

    def _create_hours_breakdown_chart(self, tech_data: List[Dict]):
        """Create stacked bar chart showing PM vs CM hours"""
        frame = self.chart_frames['hours_breakdown']

        # Sort by total hours
        sorted_data = sorted(tech_data, key=lambda x: x['total_hours'], reverse=True)

        fig = Figure(figsize=(6, 3.5), dpi=80)
        ax = fig.add_subplot(111)

        technicians = [t['technician'] for t in sorted_data]
        pm_hours = [t['pm_hours'] for t in sorted_data]
        cm_hours = [t['cm_hours'] for t in sorted_data]

        x = range(len(technicians))
        width = 0.6

        # Create stacked bars
        p1 = ax.bar(x, pm_hours, width, label='PM Hours', color='#007bff', alpha=0.8)
        p2 = ax.bar(x, cm_hours, width, bottom=pm_hours, label='CM Hours', color='#fd7e14', alpha=0.8)

        # Add total labels
        for i, (pm, cm) in enumerate(zip(pm_hours, cm_hours)):
            total = pm + cm
            ax.text(i, total, f'{total:.1f}', ha='center', va='bottom', fontsize=7, fontweight='bold')

        ax.set_xlabel('Technician', fontsize=9, fontweight='bold')
        ax.set_ylabel('Hours', fontsize=9, fontweight='bold')
        ax.set_title('Work Hours Breakdown: PM vs CM', fontsize=10, fontweight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels(technicians, rotation=45, ha='right', fontsize=8)
        ax.legend(fontsize=8)
        ax.grid(axis='y', alpha=0.3)

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)
        self.canvas_widgets.append(canvas.get_tk_widget())

    def _create_workload_distribution_chart(self, tech_data: List[Dict]):
        """Create pie chart showing workload distribution"""
        frame = self.chart_frames['workload_distribution']

        fig = Figure(figsize=(6, 3.5), dpi=80)

        # Create two subplots - one for hours, one for task count
        ax1 = fig.add_subplot(121)
        ax2 = fig.add_subplot(122)

        # Hours distribution
        technicians = [t['technician'] for t in tech_data if t['total_hours'] > 0]
        total_hours = [t['total_hours'] for t in tech_data if t['total_hours'] > 0]

        if total_hours:
            ax1.pie(total_hours, labels=technicians, autopct='%1.1f%%', startangle=90, textprops={'fontsize': 7})
            ax1.set_title('Total Hours Distribution', fontsize=10, fontweight='bold')

        # Task count distribution
        task_counts = [t['pm_count'] + t['cm_count'] for t in tech_data if (t['pm_count'] + t['cm_count']) > 0]
        tech_names = [t['technician'] for t in tech_data if (t['pm_count'] + t['cm_count']) > 0]

        if task_counts:
            ax2.pie(task_counts, labels=tech_names, autopct='%1.1f%%', startangle=90, textprops={'fontsize': 7})
            ax2.set_title('Task Count Distribution', fontsize=10, fontweight='bold')

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)
        self.canvas_widgets.append(canvas.get_tk_widget())

    def _create_trend_chart(self, tech_data: List[Dict]):
        """Create chart showing efficiency vs target"""
        frame = self.chart_frames['trend_analysis']

        fig = Figure(figsize=(6, 3.5), dpi=80)
        ax = fig.add_subplot(111)

        # Sort by technician name for consistent display
        sorted_data = sorted(tech_data, key=lambda x: x['technician'])

        technicians = [t['technician'] for t in sorted_data]
        efficiencies = [t['efficiency'] for t in sorted_data]
        x = range(len(technicians))

        # Plot efficiency line
        ax.plot(x, efficiencies, marker='o', linewidth=1.5, markersize=6, label='Actual Efficiency', color='#007bff')

        # Plot target line
        target_line = [self.TARGET_EFFICIENCY * 100] * len(technicians)
        ax.plot(x, target_line, linestyle='--', linewidth=1.5, label=f'Target ({self.TARGET_EFFICIENCY * 100}%)', color='red')

        # Fill area between efficiency and target
        ax.fill_between(x, efficiencies, target_line, alpha=0.2, color='green', where=[e >= t for e, t in zip(efficiencies, target_line)])
        ax.fill_between(x, efficiencies, target_line, alpha=0.2, color='red', where=[e < t for e, t in zip(efficiencies, target_line)])

        ax.set_xlabel('Technician', fontsize=9, fontweight='bold')
        ax.set_ylabel('Efficiency (%)', fontsize=9, fontweight='bold')
        ax.set_title('Efficiency vs Target Analysis', fontsize=10, fontweight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels(technicians, rotation=45, ha='right', fontsize=8)
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)
        self.canvas_widgets.append(canvas.get_tk_widget())

    def print_report(self):
        """Print the current efficiency report"""
        try:
            messagebox.showinfo(
                "Print Report",
                "Printing functionality:\n\n"
                "1. Use 'Export to PDF' button to create a PDF\n"
                "2. Open the PDF and print from your PDF viewer\n\n"
                "This provides better formatting and quality for printed reports."
            )
        except Exception as e:
            messagebox.showerror("Error", f"Print failed:\n{str(e)}")

    def export_to_excel(self):
        """Export report data to Excel file"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Get file path from user
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"Efficiency_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

            if not file_path:
                return

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Efficiency Report"

            # Add title
            ws['A1'] = "Technician Efficiency Report"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:J1')

            # Add date range
            ws['A2'] = f"Period: {self.start_date_var.get()} to {self.end_date_var.get()}"
            ws['A2'].font = Font(size=12)
            ws.merge_cells('A2:J2')

            # Add summary
            row = 4
            ws[f'A{row}'] = "OVERALL SUMMARY"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1

            summary_data = [
                ('Total Available Hours', self.summary_labels['total_available'].cget('text')),
                ('Total Worked Hours', self.summary_labels['total_worked'].cget('text')),
                ('Overall Efficiency', self.summary_labels['overall_efficiency'].cget('text')),
                ('PM Hours', self.summary_labels['pm_hours'].cget('text')),
                ('CM Hours', self.summary_labels['cm_hours'].cget('text')),
                ('Target Efficiency', self.summary_labels['target'].cget('text')),
            ]

            for label, value in summary_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

            # Add technician details
            row += 2
            ws[f'A{row}'] = "TECHNICIAN PERFORMANCE DETAILS"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1

            # Headers
            headers = ['Technician', 'Available Hrs', 'PM Hours', 'CM Hours', 'Total Hours',
                      'PM Count', 'CM Count', 'Efficiency %', 'vs Target', 'Status']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            row += 1

            # Data rows
            for item in self.tech_tree.get_children():
                values = self.tech_tree.item(item)['values']
                for col, value in enumerate(values, start=1):
                    ws.cell(row=row, column=col, value=value)
                row += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # Save workbook
            wb.save(file_path)
            messagebox.showinfo("Success", f"Report exported to:\n{file_path}")

        except ImportError:
            messagebox.showerror("Error", "openpyxl library not installed.\nInstall with: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed:\n{str(e)}")
            import traceback
            print(traceback.format_exc())

    def export_to_pdf(self):
        """Export report to PDF format"""
        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.enums import TA_CENTER, TA_LEFT

            # Get file path from user
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile=f"Efficiency_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            )

            if not file_path:
                return

            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=landscape(letter))
            elements = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#003366'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            elements.append(Paragraph("Technician Efficiency Report", title_style))

            # Date range
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=TA_CENTER,
                spaceAfter=20
            )
            elements.append(Paragraph(
                f"Report Period: {self.start_date_var.get()} to {self.end_date_var.get()}",
                date_style
            ))

            elements.append(Spacer(1, 0.3*inch))

            # Summary section
            elements.append(Paragraph("Overall Summary", styles['Heading2']))

            summary_data = [
                ['Metric', 'Value'],
                ['Total Available Hours', self.summary_labels['total_available'].cget('text')],
                ['Total Worked Hours', self.summary_labels['total_worked'].cget('text')],
                ['Overall Efficiency', self.summary_labels['overall_efficiency'].cget('text')],
                ['PM Hours', self.summary_labels['pm_hours'].cget('text')],
                ['CM Hours', self.summary_labels['cm_hours'].cget('text')],
                ['Target Efficiency', self.summary_labels['target'].cget('text')],
            ]

            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(summary_table)
            elements.append(Spacer(1, 0.3*inch))

            # Technician details
            elements.append(Paragraph("Technician Performance Details", styles['Heading2']))

            # Get data from treeview
            detail_data = [['Technician', 'Avail Hrs', 'PM Hrs', 'CM Hrs', 'Total Hrs',
                           'PM Cnt', 'CM Cnt', 'Efficiency', 'vs Target', 'Status']]

            for item in self.tech_tree.get_children():
                values = self.tech_tree.item(item)['values']
                detail_data.append(values)

            # Create table with adjusted column widths
            col_widths = [1.5*inch, 0.8*inch, 0.7*inch, 0.7*inch, 0.8*inch,
                         0.6*inch, 0.6*inch, 0.9*inch, 0.9*inch, 1*inch]

            detail_table = Table(detail_data, colWidths=col_widths)
            detail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))

            elements.append(detail_table)

            # Add notes
            elements.append(Spacer(1, 0.3*inch))
            elements.append(Paragraph("Report Notes", styles['Heading3']))

            notes_text = f"""
            <para>
            <b>Efficiency Calculation Methodology:</b><br/>
            • Annual Working Hours per Technician: {self.ANNUAL_HOURS_PER_TECH} hours<br/>
            • Total Technicians: {self.TOTAL_TECHNICIANS}<br/>
            • Weekly Team Availability: {self.WEEKLY_AVAILABILITY_HOURS} hours<br/>
            • Target Efficiency: {self.TARGET_EFFICIENCY * 100}%<br/>
            <br/>
            <b>Efficiency Formula:</b> (Total Worked Hours / Total Available Hours) × 100%<br/>
            <br/>
            Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
            </para>
            """

            elements.append(Paragraph(notes_text, styles['Normal']))

            # Build PDF
            doc.build(elements)
            messagebox.showinfo("Success", f"Report exported to PDF:\n{file_path}")

        except ImportError:
            messagebox.showerror(
                "Error",
                "reportlab library not installed.\n\nInstall with: pip install reportlab"
            )
        except Exception as e:
            messagebox.showerror("Error", f"PDF export failed:\n{str(e)}")
            import traceback
            print(traceback.format_exc())
